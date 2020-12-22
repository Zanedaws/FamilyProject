import flask
import sqlite3
import hashlib
import time
import math
import openpyxl as op
from tempfile import NamedTemporaryFile

app = flask.Flask(__name__, static_folder="./static")


@app.route("/genUsers/awdsfh1235dsfidofaj12k35diaso2134214dsjkahfl!")
def genUsers_get():

    user = "Kara"
    pswd = "password"

    userHash = hashlib.sha256((user + pswd).encode('utf-8')).hexdigest()

    conn = sqlite3.connect("data/database.db")
    c = conn.cursor()
    c.execute("INSERT INTO users VALUES (?,?,?)", (user, pswd, userHash))

    user = "Patrick"
    pswd = "password"

    userHash = hashlib.sha256((user + pswd).encode('utf-8')).hexdigest()

    c.execute("INSERT INTO users VALUES (?,?,?)", (user, pswd, userHash))
    conn.commit()
    conn.close()
    return flask.redirect(flask.url_for("login_get"))


@app.route('/', methods=["GET"])
def home_get():
    return flask.redirect(flask.url_for("login_get"))


@app.route('/login', methods=["POST"])
def login_post():

    userName = flask.request.form['user']
    password = flask.request.form['pass']

    userHash = hashlib.sha256((userName + password).encode('utf-8')).hexdigest()

    conn = sqlite3.connect('data/database.db')
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE userHash=?", (userHash,))
    user = c.fetchall()
    conn.close()
    data = flask.jsonify({"data": user, "hash": userHash}, 204)
    return data


@app.route('/login', methods=["GET"])
def login_get():
    return flask.render_template("home.html")


@app.route("/feed/<userHash>/<view>", methods=["GET"])
def feed_get(userHash, view):
    conn = sqlite3.connect("data/database.db")
    c = conn.cursor()
    offset = int(view) * 5
    # print(offset)
    c.execute("SELECT * FROM inputs")
    allInputs = c.fetchall()
    try:
        if len(allInputs) % 5 == 0:
            pages = len(allInputs)/5 - 1
        else:
            pages = math.floor(len(allInputs)/5)
    except:
        pages = 0
    c.execute("SELECT * FROM inputs ORDER BY timeValue DESC LIMIT 5 OFFSET ?", (offset,))
    data = c.fetchall()
    conn.close()
    # print(pages)
    data.insert(0, pages)
    data.insert(0, view)
    data.insert(0, userHash)
    return flask.render_template("feed.html", data=data)


@app.route("/profile/<userHash>", methods=["GET"])
def profile_get(userHash):
    conn = sqlite3.connect("data/database.db")
    c = conn.cursor()
    c.execute("SELECT * FROM inputs WHERE userHash=?", (userHash,))
    inputs = c.fetchall()
    conn.close()
    inputs.insert(0, userHash)
    return flask.render_template("profile.html", data=inputs)


@app.route("/input/<userHash>/<view>", methods=["GET"])
def input_get(userHash, view):
    return flask.render_template("input.html", data=[userHash, view])


@app.route("/input", methods=["POST"])
def input_post():

    note = flask.request.form['memo']
    amount = flask.request.form['amount']
    currentTime = flask.request.form['date']
    userHash = flask.request.form["userHash"]
    view = flask.request.form["view"]
    timeValue = flask.request.form["time"]

    inputHash = hashlib.sha256((str(amount) + str(note) + str(time.time())).encode('utf-8')).hexdigest()

    conn = sqlite3.connect("data/database.db")
    c = conn.cursor()

    c.execute("INSERT INTO inputs (userHash, inputHash, note, amount, inputTime, timeValue) VALUES (?,?,?,?,?,?)",
              (userHash, inputHash, note, amount, currentTime, int(timeValue)))

    conn.commit()
    conn.close()

    return flask.jsonify({"data": flask.url_for("feed_get", userHash=userHash, view=view)}, 204)


@app.route("/delete/<inputHash>", methods=["POST"])
def delete_post(inputHash):

    conn = sqlite3.connect(database="data/database.db")
    c = conn.cursor()
    c.execute("DELETE FROM inputs WHERE inputHash=?", (inputHash,))
    conn.commit()
    conn.close()

    return flask.jsonify({}, 204)


@app.route("/download", methods=["GET"])
def download_get():
    conn = sqlite3.connect(database="data/database.db")
    c = conn.cursor()
    c.execute("SELECT * FROM inputs")
    data = c.fetchall()
    conn.close()

    wb = op.Workbook()
    wb.template = False
    ws = wb.active
    ws.title = "Data"

    i = 2

    ws.cell(row=1, column=1, value="Notes")
    ws.cell(row=1, column=2, value="Miles")
    ws.cell(row=1, column=3, value="Date")

    for post in data:
        ws.cell(row=i, column=1).value = post[2]
        ws.cell(row=i, column=2).value = post[3]
        ws.cell(row=i, column=3).value = post[4]
        i += 1

    wb.save("data/Data.xlsx")
    wb.close()

    return flask.send_file("data/Data.xlsx", as_attachment=True, attachment_filename="Data.xlsx")

    ##Temp file Testing##
    # with NamedTemporaryFile() as tmp:
    #     wb.save(tmp.name)
    #     tmp.seek(0)
    #     stream = tmp.read()
    #     return flask.send_file(stream, as_attachment=True, attachment_filename="Data.xlsx")


if __name__ == "__main__":
    app.run(port="5001", host="127.0.0.1", debug=True, use_evalex=False)
